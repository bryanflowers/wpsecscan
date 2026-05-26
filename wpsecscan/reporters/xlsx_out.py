"""XLSX export — formula-injection-safe Excel workbook.

Layout:
  Sheet "Summary" — risk-score banner, grade, per-severity counts, scan meta
  One sheet per OWASP Top 10 category (only categories that have findings)
  Sheet "All findings" — flat list of every finding, sortable + filterable

Cells starting with =, +, -, @, tab, or CR are prefixed with `'` per the same
OWASP CSV-injection guideline csv_out uses.
"""
from __future__ import annotations

from pathlib import Path

from ..models import ScanReport
from .csv_out import _safe_cell


def _open_pyxl():
    """Late-import openpyxl so the rest of the package still works without it."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    return Workbook, Font, PatternFill, Alignment, get_column_letter


SEVERITY_FILL = {
    "critical": "67000D",
    "high":     "5A1816",
    "medium":   "4A3A10",
    "low":      "133246",
    "info":     "21262D",
}
SEVERITY_FG = {
    "critical": "FFD6D6",
    "high":     "FF8A85",
    "medium":   "F0C674",
    "low":      "79C0FF",
    "info":     "8B949E",
}


def write(report: ScanReport, path: Path) -> None:
    Workbook, Font, PatternFill, Alignment, get_column_letter = _open_pyxl()
    from .. import tags as _tags
    from ..risk import risk_grade, risk_label

    wb = Workbook()
    score = report.risk_score
    grade = risk_grade(score)

    # ----- Summary sheet -----
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "WPSecScan"
    ws["A1"].font = Font(size=18, bold=True)
    ws["A2"] = report.target
    ws["A2"].font = Font(size=12, color="79C0FF")
    ws["A4"] = "Scanned at"
    ws["B4"] = report.scanned_at
    ws["A5"] = "Duration (ms)"
    ws["B5"] = report.duration_ms
    ws["A6"] = "Risk score"
    ws["B6"] = f"{score}/100"
    ws["A7"] = "Risk grade"
    ws["B7"] = grade
    ws["A8"] = "Label"
    ws["B8"] = risk_label(score)

    ws["A10"] = "Severity"
    ws["B10"] = "Count"
    ws["A10"].font = Font(bold=True)
    ws["B10"].font = Font(bold=True)
    row = 11
    s = report.summary
    for sev in ("critical", "high", "medium", "low", "info"):
        ws.cell(row=row, column=1, value=sev.upper())
        ws.cell(row=row, column=2, value=s.get(sev, 0))
        ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=SEVERITY_FILL[sev])
        ws.cell(row=row, column=1).font = Font(color=SEVERITY_FG[sev], bold=True)
        row += 1
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 30

    # ----- Per-OWASP-category sheets + All-findings sheet -----
    from .. import confidence as _conf
    waf_detected = any(
        ("WAF" in (f.title or "") or "CDN detected" in (f.title or ""))
        for r in report.results if r.check_id == "waf"
        for f in r.findings
    )
    owasp_buckets: dict[str, list[tuple]] = {}  # owasp -> [(check_id, finding, ...)]
    all_rows: list[tuple] = []
    for r in report.results:
        if not r.findings:
            continue
        tg = _tags.get_tags(r.check_id) or {}
        owasp = tg.get("owasp") or "Unmapped"
        for f in r.findings:
            row_tuple = (
                _safe_cell(r.check_id),
                _safe_cell(r.check_name),
                _safe_cell(f.severity),
                _safe_cell(_conf.compute_confidence(f, r.check_id, waf_detected=waf_detected)),
                _safe_cell(f.title),
                _safe_cell(f.url),
                _safe_cell((f.evidence or "").replace("\r", " ").replace("\n", " | ")[:2000]),
                _safe_cell((f.remediation or "").replace("\r", " ").replace("\n", " | ")[:2000]),
                _safe_cell(owasp),
                _safe_cell(tg.get("attack", "")),
            )
            owasp_buckets.setdefault(owasp, []).append(row_tuple)
            all_rows.append(row_tuple)

    HEADERS = ("check_id", "check_name", "severity", "confidence", "title", "url",
               "evidence", "remediation", "owasp", "attack")

    def _write_sheet(sheet, rows):
        # Header row
        for col, h in enumerate(HEADERS, start=1):
            c = sheet.cell(row=1, column=col, value=h)
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor="22262D")
            c.font = Font(bold=True, color="E6EDF3")
        # Data rows
        for r_idx, row_tuple in enumerate(rows, start=2):
            for col, val in enumerate(row_tuple, start=1):
                cell = sheet.cell(row=r_idx, column=col, value=val)
                if col == 3:  # severity column
                    fill = SEVERITY_FILL.get(val.lower())
                    fg = SEVERITY_FG.get(val.lower())
                    if fill:
                        cell.fill = PatternFill("solid", fgColor=fill)
                        cell.font = Font(color=fg, bold=True)
                cell.alignment = Alignment(vertical="top", wrap_text=col in (5, 7, 8))
        # Reasonable column widths (after inserting confidence at col 4)
        widths = (16, 28, 10, 14, 50, 50, 70, 50, 14, 14)
        for col, w in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(col)].width = w
        # Freeze header row
        sheet.freeze_panes = "A2"

    # All-findings sheet
    if all_rows:
        ws_all = wb.create_sheet("All findings")
        _write_sheet(ws_all, all_rows)

    # Per-OWASP-category sheets — Excel forbids `: / \ ? * [ ]` in sheet names
    # and caps title length at 31 chars.
    import re as _re
    for owasp, rows in sorted(owasp_buckets.items()):
        safe_name = _re.sub(r"[:/\\?*\[\]]", "-", owasp)[:31]
        ws_o = wb.create_sheet(safe_name)
        _write_sheet(ws_o, rows)

    wb.save(str(path))
    try:
        from .. import activity as _act
        _act.emit("reporter", f"Excel: {path.name} ({path.stat().st_size // 1024} KB)")
    except (ImportError, OSError):
        pass


def render(report: ScanReport) -> bytes:
    """For symmetry with other reporters — returns the XLSX bytes."""
    import tempfile
    import os
    fd, name = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    try:
        write(report, Path(name))
        return Path(name).read_bytes()
    finally:
        try:
            os.unlink(name)
        except OSError:
            pass
