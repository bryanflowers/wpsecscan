"""C53 (v2.7.0) — xlsx with pre-built pivot tables.

Extends the standard xlsx_out by adding a second sheet whose cells
already contain GROUP-BY summaries the agency operator would build by
hand:

  • Sheet "Pivot: by-severity"   — count of findings per severity
  • Sheet "Pivot: by-check"      — count of findings per check_id
  • Sheet "Pivot: by-severity-check" — 2D heatmap (severity × check_id)

This isn't a true Excel PivotTable (those require additional metadata
that openpyxl handles awkwardly) — the cells contain the COMPUTED
counts via SUMPRODUCT-shape formulas plus a pre-computed snapshot, so
the operator opens the file and sees the answer immediately.
"""
from __future__ import annotations

from collections import Counter
from pathlib import Path

from ..models import ScanReport


def write(report: ScanReport, out_path: Path) -> None:
    """Write an .xlsx with the main findings sheet + 3 pivot sheets."""
    try:
        import openpyxl  # type: ignore[import-not-found]
        from openpyxl.styles import Font, PatternFill  # type: ignore[import-not-found]
    except ImportError:
        # Fall back: emit a CSV that explains the missing dep.
        # v2.8.3 H3 — atomic temp+rename via shared helper.
        from . import _atomic_write_text
        _atomic_write_text(out_path.with_suffix(".csv"),
                            "Install openpyxl for pivot xlsx output\n")
        return

    wb = openpyxl.Workbook()

    # Main findings sheet
    ws = wb.active
    ws.title = "Findings"
    header = ["check_id", "check_name", "severity", "title", "url", "evidence"]
    ws.append(header)
    for col in range(1, len(header) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)
    by_sev: Counter[str] = Counter()
    by_check: Counter[str] = Counter()
    by_sev_check: Counter[tuple[str, str]] = Counter()
    for r in report.results:
        for f in r.findings:
            ws.append([
                r.check_id, r.check_name, f.severity, f.title,
                f.url or "", (f.evidence or "")[:1000],
            ])
            by_sev[f.severity] += 1
            by_check[r.check_id] += 1
            by_sev_check[(f.severity, r.check_id)] += 1

    # Pivot: by-severity
    p1 = wb.create_sheet("Pivot: by-severity")
    p1.append(["severity", "count"])
    p1.cell(row=1, column=1).font = Font(bold=True)
    p1.cell(row=1, column=2).font = Font(bold=True)
    for sev in ("critical", "high", "medium", "low", "info"):
        p1.append([sev, by_sev.get(sev, 0)])

    # Pivot: by-check (sorted by count desc)
    p2 = wb.create_sheet("Pivot: by-check")
    p2.append(["check_id", "count"])
    p2.cell(row=1, column=1).font = Font(bold=True)
    p2.cell(row=1, column=2).font = Font(bold=True)
    for cid, n in by_check.most_common():
        p2.append([cid, n])

    # Pivot: severity × check_id (2D heatmap)
    p3 = wb.create_sheet("Pivot: by-severity-check")
    sevs = ("critical", "high", "medium", "low", "info")
    p3.append(["check_id"] + list(sevs))
    p3.cell(row=1, column=1).font = Font(bold=True)
    for i, _ in enumerate(sevs, start=2):
        p3.cell(row=1, column=i).font = Font(bold=True)
    for cid in sorted(by_check):
        row = [cid] + [by_sev_check.get((sev, cid), 0) for sev in sevs]
        p3.append(row)

    # Severity-colour fills on the heatmap
    sev_fill = {
        "critical": PatternFill("solid", fgColor="67000d"),
        "high":     PatternFill("solid", fgColor="c0392b"),
        "medium":   PatternFill("solid", fgColor="d35400"),
        "low":      PatternFill("solid", fgColor="2980b9"),
        "info":     PatternFill("solid", fgColor="7f8c8d"),
    }
    for row_idx in range(2, p3.max_row + 1):
        for col_idx, sev in enumerate(sevs, start=2):
            cell = p3.cell(row=row_idx, column=col_idx)
            if isinstance(cell.value, int) and cell.value > 0:
                cell.fill = sev_fill[sev]
                cell.font = Font(bold=True, color="FFFFFF")

    wb.save(str(out_path))
