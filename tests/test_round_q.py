"""Smoke tests for the round-Q 24-feature batch.

Pins inventory growth (94 checks post-round-54) and exercises each new
module's happy path. The threshold bumps as new rounds land so that an
accidental removal of a registered check trips the test.
"""
from __future__ import annotations

import json

from wpsecscan.checks import ALL_CHECKS


def test_inventory_grew_to_90_plus():
    # Round-54 brought the inventory to 94. Threshold is set to 90 so a
    # routine removal/rename of a single check doesn't trip CI, but a
    # large accidental delete does.
    assert len(ALL_CHECKS) >= 90, f"expected >=90 checks, got {len(ALL_CHECKS)}"


def test_new_check_ids_registered():
    ids = {cid for cid, _n, _f, _a in ALL_CHECKS}
    expected_new = {
        "well_known", "login_timing", "sitemap_cve_probe",   # passive
        "xxe_upload",                                          # aggressive
    }
    missing = expected_new - ids
    assert not missing, f"new round-Q checks not registered: {missing}"


def test_xxe_upload_is_marked_aggressive():
    """xxe_upload uploads to forms — must be gated by --aggressive."""
    entry = next((c for c in ALL_CHECKS if c[0] == "xxe_upload"), None)
    assert entry is not None
    assert entry[3] is True, "xxe_upload must be aggressive=True"


# ---- D1 XLSX smoke ----

def test_xlsx_reporter_writes_valid_xlsx(tmp_path):
    from wpsecscan.reporters import xlsx_out
    from wpsecscan.models import Finding, CheckResult, ScanReport
    r = ScanReport(
        target="https://example.com",
        scanned_at="2026-05-23T00:00:00Z",
        duration_ms=100,
        results=[
            CheckResult(check_id="sqli", check_name="SQL injection probes",
                        findings=[Finding(severity="high", title="=cmd|calc.exe", evidence="x")]),
        ],
    )
    p = tmp_path / "out.xlsx"
    xlsx_out.write(r, p)
    assert p.exists()
    # Reopen and verify the formula-prefixed title was neutralised
    from openpyxl import load_workbook
    wb = load_workbook(str(p))
    assert "Summary" in wb.sheetnames
    assert "All findings" in wb.sheetnames
    ws = wb["All findings"]
    # Look up title column by header so the test isn't tied to column position
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    title_col = headers.index("title") + 1
    title_cell = ws.cell(row=2, column=title_col).value
    assert title_cell.startswith("'="), f"CSV-injection-safe prefix missing: {title_cell!r}"


# ---- D3 markdown CLI smoke ----

def test_markdown_reporter_render():
    from wpsecscan.reporters import markdown as md
    from wpsecscan.models import Finding, CheckResult, ScanReport
    r = ScanReport(
        target="https://example.com",
        scanned_at="2026-05-23T00:00:00Z",
        duration_ms=0,
        results=[
            CheckResult(check_id="sqli", check_name="SQL injection probes",
                        findings=[Finding(severity="high", title="x", evidence="```nested```")]),
        ],
    )
    text = md.render(r)
    assert "# WPSecScan" in text
    assert "Risk score" in text
    # B40 (v2.8.0) — was: 3-backtick fence with ZWS hack. Now: 4-
    # backtick fence (CommonMark allows any backtick run >= 3; closer
    # must match length). Embedded literal ``` no longer needs escape.
    assert "````" in text
    # The literal embedded ``` from the evidence is preserved as-is
    # (no zero-width-space mutation).
    assert "```nested```" in text


def test_markdown_reporter_top_n_keeps_highest_severities():
    """--md-top N must keep the N most-severe findings (cross-check), preserving
    their per-check grouping. Slack-style truncation."""
    from wpsecscan.reporters import markdown as md
    from wpsecscan.models import Finding, CheckResult, ScanReport
    r = ScanReport(
        target="https://example.com",
        scanned_at="2026-05-23T00:00:00Z",
        duration_ms=0,
        results=[
            CheckResult(check_id="a", check_name="check a", findings=[
                Finding(severity="info", title="A-info"),
                Finding(severity="critical", title="A-crit"),
            ]),
            CheckResult(check_id="b", check_name="check b", findings=[
                Finding(severity="medium", title="B-med"),
                Finding(severity="high", title="B-high"),
            ]),
        ],
    )
    text = md.render(r, top_n=2)
    assert "A-crit" in text
    assert "B-high" in text
    assert "A-info" not in text
    assert "B-med" not in text


# ---- A6 Patchstack: function exists ----

def test_patchstack_function_exists_and_handles_no_token():
    from wpsecscan import db as vulndb
    assert hasattr(vulndb, "fetch_patchstack")
    out = vulndb.fetch_patchstack("")  # empty token short-circuits
    assert out == []


# ---- A7 OSV.dev helper exists ----

def test_osv_helper_module_exists():
    from wpsecscan.checks.js_libraries import _query_osv, OSV_PACKAGE_MAP
    assert "jQuery" in OSV_PACKAGE_MAP
    assert callable(_query_osv)


# ---- B1 checkpoint round-trip ----

def test_checkpoint_path_uses_safe_filename(tmp_path, monkeypatch):
    monkeypatch.setenv("WPSECSCAN_HOME", str(tmp_path))
    from wpsecscan import history as _h
    name = _h._safe_filename("https://Example.COM/blog/")
    # Just confirm the helper exists and produces something filename-safe
    assert name and "/" not in name and ":" not in name


# ---- D2 GitHub Issues: repo validation ----

def test_github_repo_validation():
    from wpsecscan.integrations import github_issues as gh
    assert gh.validate_repo("octocat/Hello-World")[0]
    assert not gh.validate_repo("not a repo")[0]
    assert not gh.validate_repo("")[0]


def test_github_create_issues_returns_summary_with_empty_token():
    from wpsecscan.integrations import github_issues as gh
    from wpsecscan.models import Finding, CheckResult, ScanReport
    r = ScanReport(
        target="https://example.com", scanned_at="t", duration_ms=0,
        results=[CheckResult(check_id="sqli", check_name="x",
                              findings=[Finding(severity="high", title="x")])],
    )
    s = gh.create_issues_for_report(r, "owner/repo", "", threshold="high")
    assert s["ok"] == 0
    assert "no token" in (s["errors"][0] if s["errors"] else "")


# ---- D4 WAF rule generator ----

def test_waf_rule_generator_exists_for_common_checks():
    from wpsecscan import waf_rules as wr
    for cid in ("exposed_files", "xmlrpc_deep", "login_throttle", "tls_headers", "users"):
        rule = wr.get_rule(cid)
        assert rule is not None, f"missing WAF rule for {cid}"
        assert rule.get("cloudflare") or rule.get("modsecurity") or rule.get("nginx")


# ---- C5 recommendation engine ----

def test_recommendation_engine_groups_findings():
    from wpsecscan import recommend as rec
    from wpsecscan.models import Finding, CheckResult, ScanReport
    r = ScanReport(
        target="https://example.com", scanned_at="t", duration_ms=0,
        results=[
            CheckResult(check_id="login_throttle", check_name="x",
                         findings=[Finding(severity="high", title="no throttle")]),
            CheckResult(check_id="login_throttle_deep", check_name="x",
                         findings=[Finding(severity="medium", title="no throttle deep")]),
        ],
    )
    recs = rec.recommendations_for(r)
    # Both findings map to the same recommendation key — should dedupe to 1
    keys = [x["key"] for x in recs]
    assert keys.count("rate-limit-login") == 1


def test_recommendation_engine_skips_info_only():
    from wpsecscan import recommend as rec
    from wpsecscan.models import Finding, CheckResult, ScanReport
    r = ScanReport(
        target="https://example.com", scanned_at="t", duration_ms=0,
        results=[
            CheckResult(check_id="login_throttle", check_name="x",
                         findings=[Finding(severity="info", title="all good")]),
        ],
    )
    assert rec.recommendations_for(r) == []


# ---- E1 custom check loader: doesn't crash when plugins dir is missing ----

def test_custom_check_loader_no_plugins_dir(tmp_path, monkeypatch):
    monkeypatch.setenv("WPSECSCAN_HOME", str(tmp_path))
    # Force re-discovery by clearing the cached flag
    import wpsecscan.checks as checks_mod
    checks_mod._CUSTOM_CHECKS_LOADED = False
    checks_mod._load_custom_checks()  # should silently no-op
    checks_mod._CUSTOM_CHECKS_LOADED = True


def test_custom_check_loader_picks_up_valid_plugin(tmp_path, monkeypatch):
    monkeypatch.setenv("WPSECSCAN_HOME", str(tmp_path))
    plug_dir = tmp_path / "plugins"
    plug_dir.mkdir()
    (plug_dir / "my_check.py").write_text("""
from wpsecscan.models import Finding
CHECK_ID = "my_test_check"
CHECK_NAME = "User test check"
IS_AGGRESSIVE = False
async def check(client, ctx):
    return [Finding(severity="info", title="user plugin fired")]
""", encoding="utf-8")
    import wpsecscan.checks as checks_mod
    checks_mod._CUSTOM_CHECKS_LOADED = False
    # Snapshot current ALL_CHECKS so we can restore + dedupe
    before = list(checks_mod.ALL_CHECKS)
    try:
        checks_mod._load_custom_checks()
        ids = {cid for cid, _n, _f, _a in checks_mod.ALL_CHECKS}
        assert "my_test_check" in ids
    finally:
        # Restore so subsequent tests aren't polluted
        checks_mod.ALL_CHECKS[:] = before
        checks_mod._CUSTOM_CHECKS_LOADED = True


# ---- C1 disabled-checks file is honored ----

def test_disabled_checks_filter(tmp_path, monkeypatch):
    monkeypatch.setenv("WPSECSCAN_HOME", str(tmp_path))
    # Write a disable list
    (tmp_path / "disabled_checks.json").write_text(json.dumps(["waf", "core_version"]), encoding="utf-8")
    from wpsecscan.checks import select_checks
    selected = select_checks(aggressive=False, authenticated_enabled=False)
    ids = {cid for cid, _n, _f in selected}
    assert "waf" not in ids
    assert "core_version" not in ids
    # Other passive checks should still be there
    assert "users" in ids


# ---- All round-Q tags + compliance coverage ----

def test_all_round_q_checks_have_tags():
    from wpsecscan import tags as _tags
    _tags.reset_cache()
    for cid in ("well_known", "login_timing", "sitemap_cve_probe", "xxe_upload"):
        assert _tags.get_tags(cid) is not None, f"{cid} missing tag entry"
        assert _tags.get_compliance(cid) is not None, f"{cid} missing compliance entry"
